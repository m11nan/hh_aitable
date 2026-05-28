import json
import math
import os
import re
import sys
import time
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from rich.prompt import Confirm

from utils.config import get_config
from utils.logger import setup_logger
from utils.progress_helper import make_progress

logger = setup_logger()

AI_COLUMNS = [
    "AI_Score",
    "AI_Verdict",
    "AI_Status",
    "AI_Error",
]


def _ensure_cuda_paths():
    """Добавляет пути к CUDA DLL в PATH для работы llama.cpp на GPU."""
    nvidia_base = os.path.join(
        os.path.dirname(os.path.dirname(sys.executable)),
        "Lib",
        "site-packages",
        "nvidia",
    )
    cuda_dirs = [
        os.path.join(nvidia_base, "cuda_runtime", "bin"),
        os.path.join(nvidia_base, "cublas", "bin"),
        os.path.join(nvidia_base, "cuda_nvrtc", "bin"),
    ]
    paths = []
    for d in cuda_dirs:
        if os.path.exists(d):
            paths.append(d)
            os.add_dll_directory(d)
    if paths:
        os.environ["PATH"] = os.pathsep.join(paths) + os.pathsep + os.environ.get("PATH", "")

    cuda_path = os.environ.get("CUDA_PATH", "")
    if cuda_path and not os.path.exists(os.path.join(cuda_path, "bin")):
        for ver in ["v13.2", "v12.6"]:
            test = rf"C:\Program Files\NVIDIA GPU Computing Toolkit\CUDA\{ver}"
            if os.path.exists(os.path.join(test, "bin")):
                os.environ["CUDA_PATH"] = test
                break


class AIAnalyzer:
    """Запускает локальную LLM для анализа вакансий и записывает результат в Excel."""

    def __init__(self):
        """Загружает настройки модели и промптов из конфига."""
        self.model_path = get_config().get("ai_settings.model.path", "models/model.gguf")
        self.n_gpu_layers = get_config().get("ai_settings.model.n_gpu_layers", -1)
        self.n_ctx = get_config().get("ai_settings.model.n_ctx", 8192)
        self.cache_type = get_config().get("ai_settings.model.cache_type", "q8_0")
        self.flash_attn = get_config().get("ai_settings.model.flash_attn", True)
        self.max_tokens = get_config().get("ai_settings.inference.max_tokens", 1500)
        self.temperature = get_config().get("ai_settings.inference.temperature", 0.1)
        self.top_p = get_config().get("ai_settings.inference.top_p", 0.95)
        self.delay_between = get_config().get("ai_settings.inference.delay_between", 0.5)
        self.batch_save_interval = get_config().get("ai_settings.inference.batch_save_interval", 10)
        self.competition_threshold = get_config().get(
            "ai_settings.calculation.competition_threshold", 200
        )
        self.reviews_threshold = get_config().get("ai_settings.calculation.reviews_threshold", 100)
        self.system_prompt = get_config().get(
            "ai_settings.prompt.system",
            "Ты — строгий аналитик вакансий.",
        )
        self.user_prompt_template = get_config().get("ai_settings.prompt.user", "")

        self._llm: Any = None
        self._llm_available: bool | None = None
        self._has_gpu: bool | None = None

    def _init_backend(self):
        """Проверяет доступность llama-cpp-python и GPU, вызывает настройку CUDA."""
        if self._llm_available is not None:
            return

        _ensure_cuda_paths()

        try:
            from llama_cpp import Llama as LlamaCls
            from llama_cpp import llama_supports_gpu_offload

            self._llm_cls = LlamaCls
            self._llm_available = True
            self._has_gpu = llama_supports_gpu_offload()
        except ImportError:
            self._llm_available = False
            self._has_gpu = False

        if not self._llm_available:
            raise RuntimeError(
                "llama-cpp-python не установлен. Выполни: uv pip install llama-cpp-python"
            )
        if not self._has_gpu:
            logger.warning("GPU offload не поддерживается. Работаем на CPU — будет медленно.")

    def _get_model(self):
        """Загружает и возвращает модель llama.cpp (ленивая инициализация).

        Returns:
            Экземпляр Llama.
        """
        self._init_backend()
        if self._llm is None:
            if not os.path.exists(self.model_path):
                raise FileNotFoundError(
                    f"Модель не найдена: {self.model_path}. "
                    f"Укажи правильный путь в ai_settings.model.path"
                )
            logger.info(f"Загрузка модели: {self.model_path}")
            logger.info(
                f"GPU слоёв: {self.n_gpu_layers}, контекст: {self.n_ctx}, "
                f"KV cache: {self.cache_type}, flash_attn: {self.flash_attn}"
            )
            self._llm = self._llm_cls(
                model_path=self.model_path,
                n_gpu_layers=self.n_gpu_layers,
                n_ctx=self.n_ctx,
                cache_type=self.cache_type,
                flash_attn=self.flash_attn,
                verbose=False,
            )
            logger.info("Модель загружена")
        return self._llm

    def _calculate_trust_factor(self, rating: float | None, reviews: int | None) -> float:
        """Считает фактор доверия к компании на основе рейтинга и отзывов.

        Args:
            rating: Рейтинг компании (0-5) или None.
            reviews: Количество отзывов или None.

        Returns:
            Число от 0.1 до 1.0.
        """
        if not reviews or reviews <= 0:
            return 0.1
        if not rating:
            return 0.1
        return (rating / 5.0) * min(reviews / self.reviews_threshold, 1.0)

    def _calculate_competition_factor(self, total_responses: int | None) -> float:
        """Считает фактор конкуренции на основе количества откликов.

        Args:
            total_responses: Общее число откликов или None.

        Returns:
            Число от 0.0 до 1.0 (1.0 = нет конкуренции).
        """
        if not total_responses or total_responses <= 0:
            return 1.0
        return 1.0 - min(total_responses / self.competition_threshold, 1.0)

    def _build_prompt(self, row: dict[str, Any]) -> str:
        """Собирает промпт для LLM из строки данных.

        Args:
            row: Словарь с данными вакансии из Excel.

        Returns:
            Готовый промпт.
        """
        rating = row.get("Рейтинг")
        if rating is None or (isinstance(rating, float) and math.isnan(rating)):
            rating = None
        elif isinstance(rating, str) and rating == "N/A":
            rating = None
        reviews = row.get("Отзывы")
        reviews_int: int | None
        if reviews is None:
            reviews_int = None
        elif isinstance(reviews, str) and reviews == "N/A":
            reviews_int = None
        elif isinstance(reviews, float) and (math.isnan(reviews) or math.isinf(reviews)):
            reviews_int = None
        else:
            reviews_int = int(reviews)
        total_responses = row.get("Отклики_Всего", 0)
        if total_responses is None or (
            isinstance(total_responses, float) and math.isnan(total_responses)
        ):
            total_responses = 0
        total_responses_int: int | None = (
            int(total_responses) if not isinstance(total_responses, str) else None
        )
        rating_float: float | None = float(rating) if rating is not None else None
        trust_factor = self._calculate_trust_factor(rating_float, reviews_int)
        competition_factor = self._calculate_competition_factor(total_responses_int)

        salary_from = row.get("ЗП MIN", "-")
        salary_to = row.get("ЗП MAX", "-")
        currency = row.get("Валюта", "")

        vacancy_fields = {
            "name": row.get("Вакансия"),
            "company": row.get("Компания"),
            "description": row.get("Описание"),
            "salary_from": salary_from,
            "salary_to": salary_to,
            "currency": currency,
            "city": row.get("Локация"),
            "employment_type": row.get("Тип ТД"),
            "is_trusted": row.get("Верификация"),
            "is_accredited": row.get("Аккредитация"),
            "total_responses": total_responses,
        }

        filled = self.user_prompt_template.format(
            trust_factor=trust_factor,
            competition_factor=competition_factor,
            is_trusted=row.get("Верификация", "Нет"),
            is_accredited=row.get("Аккредитация", "Нет"),
            employment_type=row.get("Тип ТД", "Неизвестно"),
            salary_type=row.get("Тип ЗП", "Не указана"),
            salary_from=salary_from,
            salary_to=salary_to,
            currency=currency or "",
            total_responses=total_responses or 0,
            vacancy_json=json.dumps(vacancy_fields, ensure_ascii=False, indent=2),
        )
        return filled

    def _sanitize_parsed(self, parsed: dict[str, Any]) -> dict[str, Any]:
        """Приводит overall_score к int, убирает NaN.

        Args:
            parsed: Распарсенный ответ LLM.

        Returns:
            Очищенный словарь.
        """
        if "overall_score" in parsed:
            try:
                parsed["overall_score"] = int(float(parsed["overall_score"]))
            except (ValueError, TypeError):
                parsed["overall_score"] = 0
        return parsed

    def _parse_llm_response(self, text: str) -> dict[str, Any] | None:
        """Парсит ответ LLM в JSON с 4 уровнями fallback.

        Пытается распарсить как:
        1. Чистый JSON
        2. JSON внутри ```json ... ```
        3. Первый {...} блок
        4. JSON с починенными кавычками, запятыми, bool

        Args:
            text: Сырой ответ LLM.

        Returns:
            Словарь с ключами overall_score, explanation, is_shitty или None.
        """
        stripped = text.strip()

        try:
            return self._sanitize_parsed(json.loads(stripped))
        except json.JSONDecodeError:
            pass

        cleaned = re.sub(r"```(?:json)?\s*", "", stripped).strip()
        if cleaned != stripped:
            try:
                return self._sanitize_parsed(json.loads(cleaned))
            except json.JSONDecodeError:
                pass

        brace_match = re.search(r"\{.*\}", cleaned, re.DOTALL)
        if brace_match:
            try:
                return self._sanitize_parsed(json.loads(brace_match.group()))
            except json.JSONDecodeError:
                pass

        fixed = cleaned
        fixed = re.sub(r"(?<!\\)'", '"', fixed)
        fixed = re.sub(r",\s*}", "}", fixed)
        fixed = re.sub(r",\s*]", "]", fixed)
        fixed = re.sub(r"True|False", lambda m: m.group().lower(), fixed)
        try:
            return self._sanitize_parsed(json.loads(fixed))
        except json.JSONDecodeError:
            pass

        return None

    def _default_error_result(self, error_msg: str) -> dict[str, Any]:
        """Формирует результат-заглушку при ошибке анализа.

        Args:
            error_msg: Текст ошибки.

        Returns:
            Словарь с нулевым score и сообщением об ошибке.
        """
        return {
            "overall_score": 0,
            "explanation": f"Ошибка анализа: {error_msg}",
            "is_shitty": True,
        }

    def _save_excel(self, df: pd.DataFrame, excel_path: str):
        """Сохраняет DataFrame в Excel без потери других листов.

        Args:
            df: Данные для сохранения.
            excel_path: Путь к файлу.
        """
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            if "Вакансии" in wb.sheetnames:
                del wb["Вакансии"]
            with pd.ExcelWriter(
                excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                df.to_excel(writer, index=False, sheet_name="Вакансии")
            wb.close()
        else:
            with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, index=False, sheet_name="Вакансии")

    def analyze(self, excel_path: str):
        """Запускает AI-анализ вакансий из Excel-файла.

        Анализирует только строки с AI_Status = None или error.
        Результаты записывает в колонки AI_Score, AI_Verdict, AI_Status, AI_Error.

        Args:
            excel_path: Путь к Excel-файлу с вакансиями.
        """
        if not os.path.exists(excel_path):
            logger.error(f"Excel файл не найден: {excel_path}")
            return

        df = pd.read_excel(excel_path, engine="openpyxl")

        if "Ссылка" not in df.columns:
            logger.warning("В Excel нет колонки 'Ссылка'. Невозможно определить вакансии.")
            return

        for col in AI_COLUMNS:
            if col not in df.columns:
                df[col] = None
            df[col] = df[col].astype(object)

        needs_analysis = df["AI_Status"].isna() | (df["AI_Status"] == "error")
        rows_to_analyze = df[needs_analysis].copy()

        if rows_to_analyze.empty:
            logger.info("Все вакансии уже проанализированы.")
            return

        logger.info(f"Найдено {len(rows_to_analyze)} вакансий для AI-анализа (из {len(df)}).")

        if not Confirm.ask(
            f"[bold yellow]Запустить AI-анализ для {len(rows_to_analyze)} вакансий?[/]",
            default=False,
        ):
            logger.info("AI-анализ отменён пользователем.")
            return

        self._init_backend()
        llm = self._get_model()
        done_count = 0

        with make_progress(
            "[bold cyan]AI-анализ вакансий",
            total=len(rows_to_analyze),
            text_style="cyan",
            bar_style="cyan",
        ) as prog:
            task = prog.add_task("Анализ...", total=len(rows_to_analyze))

            for idx, (orig_idx, row) in enumerate(rows_to_analyze.iterrows()):
                try:
                    prompt = self._build_prompt(row.to_dict())
                    response = llm.create_chat_completion(
                        messages=[
                            {"role": "system", "content": self.system_prompt},
                            {"role": "user", "content": prompt},
                        ],
                        max_tokens=self.max_tokens,
                        temperature=self.temperature,
                        top_p=self.top_p,
                    )

                    raw_output = (
                        response.get("choices", [{}])[0].get("message", {}).get("content", "")
                    )

                    parsed = self._parse_llm_response(raw_output)

                    if parsed is None:
                        raise ValueError(
                            f"Не удалось распарсить ответ LLM. Сырой ответ: {raw_output[:200]}"
                        )

                    score = parsed.get("overall_score", 0)
                    try:
                        score = int(float(score))
                    except (ValueError, TypeError):
                        score = 0
                    if math.isinf(score) or math.isnan(score):
                        score = 0

                    df.at[orig_idx, "AI_Score"] = score
                    df.at[orig_idx, "AI_Verdict"] = parsed.get("explanation", "")
                    df.at[orig_idx, "AI_Status"] = "ok"
                    df.at[orig_idx, "AI_Error"] = None

                    done_count += 1
                    time.sleep(self.delay_between)

                except Exception as e:
                    logger.error(f"Ошибка AI-анализа вакансии {row.get('Вакансия', '?')}: {e}")
                    error_info = self._default_error_result(str(e))
                    df.at[orig_idx, "AI_Score"] = error_info["overall_score"]
                    df.at[orig_idx, "AI_Verdict"] = error_info["explanation"]
                    df.at[orig_idx, "AI_Status"] = "error"
                    df.at[orig_idx, "AI_Error"] = str(e)

                prog.update(task, advance=1)

                if done_count > 0 and done_count % self.batch_save_interval == 0:
                    self._save_excel(df, excel_path)
                    logger.info(f"Промежуточное сохранение: {done_count}/{len(rows_to_analyze)}")

        self._save_excel(df, excel_path)

        ok_count = (df["AI_Status"] == "ok").sum()
        err_count = (df["AI_Status"] == "error").sum()
        logger.info(
            f"AI-анализ завершён. Успешно: {ok_count}, Ошибок: {err_count}. Файл: {excel_path}"
        )
